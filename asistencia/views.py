from __future__ import annotations

import json
from datetime import timedelta

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, LogoutView
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, FormView, TemplateView
from django.db import IntegrityError
from django.db.models import DurationField, Sum, Value
from django.db.models.functions import Coalesce

from .forms import AreaForm, AusenciaProgramadaForm, EmpleadoCreationForm, HorarioForm, IpOficinaAutorizadaForm, JustificacionForm, DiaFeriadoForm, MetaHorasPracticanteForm
from .models import (
    CustomUser,
    Horario,
    IpOficinaAutorizada,
    Justificacion,
    RegistroAsistencia,
    Area,
    AusenciaProgramada,
    ConfiguracionGPS,
    DiaFeriado,
    MetaHorasPracticante,
    RecuperacionDia,
)
from .utils import calcular_horas_netas, obtener_ip_cliente, validar_ubicacion_gps


def obtener_redirect_por_rol(usuario: CustomUser) -> str:
    if usuario.rol == CustomUser.ROL_ADMIN:
        return "admin_dashboard"
    if usuario.rol == CustomUser.ROL_RRHH:
        return "dashboard"
    if usuario.rol == CustomUser.ROL_SUPERVISOR:
        return "panel_control"
    return "dashboard"


def formatear_duracion(duracion: timedelta | None) -> str:
    if not duracion:
        return "0:00"
    total_minutos = int(duracion.total_seconds() // 60)
    horas = total_minutos // 60
    minutos = total_minutos % 60
    return f"{horas}:{minutos:02d}"


def mensajes_validation_error(error: ValidationError) -> str:
    if hasattr(error, "message_dict"):
        errores = []
        for field_errors in error.message_dict.values():
            errores.extend(field_errors)
        if errores:
            return " | ".join(errores)
    if getattr(error, "messages", None):
        return " | ".join(error.messages)
    return "Datos inválidos"


def requiere_validacion_gps(usuario: CustomUser) -> bool:
    return not usuario.permite_remoto


def validar_gps_para_marcacion(usuario: CustomUser, latitud, longitud) -> str | None:
    """Devuelve mensaje de error o None si la ubicación es válida."""
    if not requiere_validacion_gps(usuario):
        return None
    if not latitud or not longitud:
        return "Debes compartir tu ubicación antes de marcar entrada."
    config_gps = ConfiguracionGPS.obtener_configuracion_activa()
    if not config_gps:
        return "No hay ubicación de oficina configurada. Contacta a administración."
    resultado = validar_ubicacion_gps(
        latitud,
        longitud,
        config_gps.latitud,
        config_gps.longitud,
        config_gps.radio_permitido_metros,
    )
    if not resultado["valido"]:
        distancia = resultado.get("distancia")
        extra = f" ({distancia}m)" if distancia is not None else ""
        return f"Debes estar en la oficina para marcar entrada{extra}. Comparte tu ubicación nuevamente."
    return None


def obtener_recuperacion_pendiente(usuario: CustomUser):
    return RecuperacionDia.objects.filter(
        empleado=usuario,
        estado=RecuperacionDia.ESTADO_PENDIENTE,
    ).order_by("fecha_falta").first()


def validar_horario_para_marcacion(usuario: CustomUser, fecha) -> str | None:
    """Valida horario de practicante/empleado antes de marcar entrada."""
    if not usuario.horario:
        return "No tienes horario asignado. Solicita a RRHH que configure tu turno."
    if not usuario.horario.es_laborable(fecha):
        # Permitir marcar si tiene recuperaciones pendientes (ej. sábado)
        if obtener_recuperacion_pendiente(usuario):
            return None  # Permitir marcar para recuperar
        return "Hoy no es día laborable según tu horario asignado."
    return None


def validar_dia_sin_permiso_ni_feriado(usuario: CustomUser, fecha) -> str | None:
    if DiaFeriado.objects.filter(fecha=fecha).exists():
        return "Hoy es feriado. No debes marcar asistencia."
    if AusenciaProgramada.objects.filter(
        empleado=usuario,
        estado=AusenciaProgramada.ESTADO_APROBADA,
        fecha_inicio__lte=fecha,
        fecha_fin__gte=fecha,
    ).exists():
        return "Tienes permiso o ausencia aprobada para hoy."
    return None


class CustomLoginView(LoginView):
    def get_success_url(self):
        return reverse(obtener_redirect_por_rol(self.request.user))


class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.rol == CustomUser.ROL_ADMIN

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            return redirect(obtener_redirect_por_rol(self.request.user))
        return super().handle_no_permission()


class AdminOrRRHHRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_authenticated and self.request.user.rol in (
            CustomUser.ROL_ADMIN,
            CustomUser.ROL_RRHH,
        )

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            return redirect(obtener_redirect_por_rol(self.request.user))
        return super().handle_no_permission()


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "asistencia/dashboard.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol not in [CustomUser.ROL_EMPLEADO, CustomUser.ROL_RRHH]:
            return redirect(obtener_redirect_por_rol(request.user))
        return super().dispatch(request, *args, **kwargs)

    def get_template_names(self):
        if self.request.user.rol == CustomUser.ROL_RRHH:
            return ["asistencia/dashboard_rrhh.html"]
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        usuario = self.request.user
        hoy = timezone.localdate()
        registro_hoy = RegistroAsistencia.objects.filter(empleado=usuario, fecha=hoy).first()
        ultimos = RegistroAsistencia.objects.filter(empleado=usuario).order_by("-fecha")[:5]
        ip_actual = obtener_ip_cliente(self.request)
        ip_valida = True
        if not usuario.permite_remoto:
            ip_valida = IpOficinaAutorizada.objects.filter(ip_publica=ip_actual, activa=True).exists()
        horario = usuario.horario
        context.update(
            {
                "registro_hoy": registro_hoy,
                "ultimos_registros": ultimos,
                "ip_actual": ip_actual,
                "ip_valida": ip_valida,
                "horario": horario,
                "dia_laborable": horario.es_laborable(hoy) if horario else None,
                "requiere_gps": requiere_validacion_gps(usuario),
                "limite_entrada": horario.hora_entrada_con_tolerancia() if horario else None,
            }
        )
        return context


class PanelControlView(LoginRequiredMixin, TemplateView):
    template_name = "asistencia/panel_control.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoy = timezone.localdate()
        empleados = CustomUser.objects.filter(
            is_active=True,
        ).select_related('area')
        asistencias_hoy = RegistroAsistencia.objects.filter(fecha=hoy)
        
        empleados_lista = []
        for empleado in empleados:
            registro_hoy = asistencias_hoy.filter(empleado=empleado).first()
            empleado.estado_hoy = (
                "Presente"
                if registro_hoy and registro_hoy.estado == RegistroAsistencia.ESTADO_A_TIEMPO
                else "Tardanza"
                if registro_hoy and registro_hoy.estado == RegistroAsistencia.ESTADO_TARDANZA
                else "Falta"
                if registro_hoy and registro_hoy.estado == RegistroAsistencia.ESTADO_FALTA
                else "No marcado"
            )
            empleado.registro_hoy = registro_hoy
            empleados_lista.append(empleado)
        
        context.update(
            {
                "total_empleados": empleados.count(),
                "presentes": asistencias_hoy.exclude(estado=RegistroAsistencia.ESTADO_FALTA).count(),
                "tardanzas": asistencias_hoy.filter(estado=RegistroAsistencia.ESTADO_TARDANZA).count(),
                "faltas": asistencias_hoy.filter(estado=RegistroAsistencia.ESTADO_FALTA).count(),
                "empleados": empleados_lista,
                "empleado_form": EmpleadoCreationForm(),
                "areas": Area.objects.all().order_by("nombre"),
            }
        )
        return context


class AdminDashboardView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    template_name = "asistencia/admin_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ip_actual = obtener_ip_cliente(self.request)
        ip_autorizada = IpOficinaAutorizada.objects.filter(ip_publica=ip_actual, activa=True).exists()
        hoy = timezone.localdate()
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        inicio_mes = hoy.replace(day=1)

        # Obtener todos los empleados activos
        empleados = CustomUser.objects.filter(is_active=True)
        registros = RegistroAsistencia.objects.filter(empleado__in=empleados)

        total_hoy = registros.filter(fecha=hoy).aggregate(
            total=Coalesce(Sum("horas_netas_trabajadas"), Value(timedelta(0), output_field=DurationField()))
        )["total"]
        total_semana = registros.filter(fecha__range=(inicio_semana, hoy)).aggregate(
            total=Coalesce(Sum("horas_netas_trabajadas"), Value(timedelta(0), output_field=DurationField()))
        )["total"]
        total_mes = registros.filter(fecha__range=(inicio_mes, hoy)).aggregate(
            total=Coalesce(Sum("horas_netas_trabajadas"), Value(timedelta(0), output_field=DurationField()))
        )["total"]

        totales_hoy = {
            item["empleado_id"]: item["total"]
            for item in registros.filter(fecha=hoy)
            .values("empleado_id")
            .annotate(total=Coalesce(Sum("horas_netas_trabajadas"), Value(timedelta(0), output_field=DurationField())))
        }
        totales_semana = {
            item["empleado_id"]: item["total"]
            for item in registros.filter(fecha__range=(inicio_semana, hoy))
            .values("empleado_id")
            .annotate(total=Coalesce(Sum("horas_netas_trabajadas"), Value(timedelta(0), output_field=DurationField())))
        }
        totales_mes = {
            item["empleado_id"]: item["total"]
            for item in registros.filter(fecha__range=(inicio_mes, hoy))
            .values("empleado_id")
            .annotate(total=Coalesce(Sum("horas_netas_trabajadas"), Value(timedelta(0), output_field=DurationField())))
        }

        horas_por_empleado = []
        for empleado in empleados.order_by("last_name", "first_name", "username"):
            nombre = empleado.get_full_name().strip() or empleado.username
            horas_por_empleado.append(
                {
                    "nombre": nombre,
                    "hoy": formatear_duracion(totales_hoy.get(empleado.id)),
                    "semana": formatear_duracion(totales_semana.get(empleado.id)),
                    "mes": formatear_duracion(totales_mes.get(empleado.id)),
                }
            )

        asistencias_hoy = RegistroAsistencia.objects.filter(fecha=hoy)
        
        # Actividades recientes de hoy (últimas 10)
        actividades_recientes = asistencias_hoy.filter(hora_entrada__isnull=False).select_related('empleado').order_by('-hora_entrada')[:10]
        
        context.update(
            {
                "ip_actual": ip_actual,
                "ip_autorizada": ip_autorizada,
                "total_empleados": empleados.count(),
                "presentes": asistencias_hoy.exclude(estado=RegistroAsistencia.ESTADO_FALTA).count(),
                "tardanzas": asistencias_hoy.filter(estado=RegistroAsistencia.ESTADO_TARDANZA).count(),
                "faltas": asistencias_hoy.filter(estado=RegistroAsistencia.ESTADO_FALTA).count(),
                "horas_hoy": formatear_duracion(total_hoy),
                "horas_semana": formatear_duracion(total_semana),
                "horas_mes": formatear_duracion(total_mes),
                "horas_por_empleado": horas_por_empleado,
                "horarios": Horario.objects.all().order_by("nombre"),
                "inicio_semana": inicio_semana,
                "inicio_mes": inicio_mes,
                "actividades_recientes": actividades_recientes,
            }
        )
        return context


@login_required
@require_POST
def autorizar_ip_actual(request):
    if request.user.rol != CustomUser.ROL_ADMIN:
        return redirect(obtener_redirect_por_rol(request.user))
    ip_actual = obtener_ip_cliente(request)
    nombre_sede = f"Auto {timezone.localdate()}"
    IpOficinaAutorizada.objects.update_or_create(
        ip_publica=ip_actual,
        defaults={"nombre_sede": nombre_sede, "activa": True},
    )
    return redirect("admin_dashboard")


class ReporteAsistenciasView(LoginRequiredMixin, TemplateView):
    template_name = "asistencia/reporte.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.rol not in [
            CustomUser.ROL_ADMIN,
            CustomUser.ROL_RRHH,
            CustomUser.ROL_SUPERVISOR,
        ]:
            return redirect(obtener_redirect_por_rol(request.user))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registros = RegistroAsistencia.objects.all()
        
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        empleado_id = self.request.GET.get('empleado')
        area_id = self.request.GET.get('area')
        
        if fecha_inicio:
            registros = registros.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            registros = registros.filter(fecha__lte=fecha_fin)
        if empleado_id:
            registros = registros.filter(empleado_id=empleado_id)
        if area_id:
            registros = registros.filter(empleado__area_id=area_id)
            
        # Mostrar agrupado/ordenado por empleado primero (Nombre → Apellido),
        # luego por fecha/hora para mantener un orden estable.
        context["registros"] = registros.order_by(
            "empleado__first_name",
            "empleado__last_name",
            "-fecha",
            "-hora_entrada",
        )
        context['empleados_list'] = CustomUser.objects.filter(rol=CustomUser.ROL_EMPLEADO, is_active=True).order_by('first_name', 'last_name')
        context['areas_list'] = Area.objects.all().order_by('nombre')
        return context



class JustificacionesPendientesView(LoginRequiredMixin, TemplateView):
    template_name = "asistencia/justificaciones_pendientes.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.rol in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
            context["justificaciones"] = Justificacion.objects.filter(aprobada=False)
        else:
            context["justificaciones"] = Justificacion.objects.filter(
                aprobada=False, 
                asistencia__empleado__supervisor=self.request.user
            )
        return context

@login_required
@require_POST
def procesar_justificacion(request, justificacion_id: int):
    accion = request.POST.get("accion")
    justificacion = get_object_or_404(Justificacion, id=justificacion_id)
    
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_SUPERVISOR, CustomUser.ROL_RRHH]:
        return JsonResponse({"status": "error", "message": "No autorizado"}, status=403)
        
    if accion == "aprobar":
        justificacion.aprobada = True
        justificacion.save()
        # Opcional: Actualizar el estado de la asistencia a 'a_tiempo' o 'justificado'
        return JsonResponse({"status": "ok", "message": "Justificación aprobada"})
    elif accion == "rechazar":
        justificacion.delete()
        return JsonResponse({"status": "ok", "message": "Justificación rechazada"})
        
    return JsonResponse({"status": "error", "message": "Acción inválida"}, status=400)


@login_required
@require_POST
def eliminar_empleado(request, empleado_id: int):
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    empleado = get_object_or_404(CustomUser, pk=empleado_id)

    if empleado == request.user:
        return JsonResponse({'success': False, 'message': 'No puedes eliminar tu propio usuario.'}, status=403)

    if empleado.rol == CustomUser.ROL_ADMIN:
        return JsonResponse({'success': False, 'message': 'No se puede eliminar un administrador.'}, status=403)

    if request.user.rol == CustomUser.ROL_RRHH and empleado.rol != CustomUser.ROL_EMPLEADO:
        return JsonResponse({'success': False, 'message': 'RRHH solo puede eliminar empleados.'}, status=403)

    empleado.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'message': 'Empleado eliminado correctamente'})

    return redirect('panel_control')


class EmpleadosView(LoginRequiredMixin, AdminOrRRHHRequiredMixin, FormView):
    template_name = "asistencia/empleados.html"
    form_class = EmpleadoCreationForm
    success_url = "/empleados/"

    def form_valid(self, form):
        # Guardamos el empleado creado de forma tradicional
        empleado = form.save()
        
        # Detectamos si la petición es asíncrona (Fetch JavaScript)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Empleado creado correctamente'
            })
            
        return super().form_valid(form)

    def form_invalid(self, form):
        # Si el formulario tiene errores (DNI duplicado, contraseñas no coinciden, etc.)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            # Estructuramos los errores de forma legible para enviárselos al Frontend
            errores_lista = []
            for campo, errores in form.errors.items():
                for error in errores:
                    errores_lista.append(f"{campo.capitalize()}: {error}")
            
            mensaje_error = " | ".join(errores_lista)
            
            return JsonResponse({
                'success': False,
                'message': mensaje_error
            }, status=200) # Respondemos 200 para capturarlo limpiamente en el .then()
            
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["empleados"] = CustomUser.objects.filter(rol=CustomUser.ROL_EMPLEADO).order_by(
            "last_name",
            "first_name",
            "username",
        )
        context["areas"] = Area.objects.all().order_by("nombre")
        return context


class ConfigIpView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    template_name = "asistencia/config_ip.html"
    form_class = IpOficinaAutorizadaForm
    success_url = "/config-ip/"


class JustificacionCreateView(LoginRequiredMixin, CreateView):
    template_name = "asistencia/justificacion_form.html"
    form_class = JustificacionForm

    def dispatch(self, request, *args, **kwargs):
        self.asistencia = get_object_or_404(RegistroAsistencia, pk=kwargs["asistencia_id"])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.asistencia = self.asistencia
        return super().form_valid(form)

    def get_success_url(self):
        return "/dashboard/"


@login_required
@require_POST
def validar_gps(request):
    """
    Valida las coordenadas GPS del usuario contra la ubicación de la oficina.
    
    Espera JSON con:
    - latitud: float
    - longitud: float
    - precisión: float (opcional)
    
    Si latitud y longitud son 0, devuelve la configuración actual en lugar de validar.
    
    Retorna JSON con:
    - valido: bool (cuando se valida)
    - distancia: float (cuando se valida)
    - mensaje: str
    - config: dict (cuando se solicita configuración)
    """
    import json
    
    try:
        datos = json.loads(request.body)
        latitud_usuario = datos.get('latitud')
        longitud_usuario = datos.get('longitud')
        precision_usuario = datos.get('precisión')
        
        # Obtener configuración GPS activa
        config_gps = ConfiguracionGPS.obtener_configuracion_activa()
        
        # Si latitud y longitud son 0, solo devolver la configuración
        if (latitud_usuario == 0 or latitud_usuario is None) and (longitud_usuario == 0 or longitud_usuario is None):
            if config_gps:
                return JsonResponse({
                    'config': {
                        'nombre': config_gps.nombre,
                        'latitud': float(config_gps.latitud),
                        'longitud': float(config_gps.longitud),
                        'radio': config_gps.radio_permitido_metros,
                        'radio_permitido_metros': config_gps.radio_permitido_metros,
                        'activa': config_gps.activa
                    }
                })
            else:
                return JsonResponse({
                    'config': None,
                    'mensaje': 'No hay configuración GPS disponible'
                })
        
        # Validar coordenadas de usuario
        if not latitud_usuario or not longitud_usuario:
            return JsonResponse({
                'valido': False,
                'mensaje': 'Coordenadas GPS inválidas o no disponibles'
            }, status=400)
        
        # Verificar que haya configuración
        if not config_gps:
            return JsonResponse({
                'valido': False,
                'mensaje': 'Configuración de ubicación de oficina no disponible'
            }, status=500)
        
        # Validar ubicación
        resultado = validar_ubicacion_gps(
            latitud_usuario,
            longitud_usuario,
            config_gps.latitud,
            config_gps.longitud,
            config_gps.radio_permitido_metros
        )
        
        return JsonResponse({
            'valido': resultado['valido'],
            'distancia': resultado['distancia'],
            'mensaje': resultado['mensaje'],
            'precision': precision_usuario,
            'radio_permitido': config_gps.radio_permitido_metros,
            'config': {
                'nombre': config_gps.nombre,
                'latitud': float(config_gps.latitud),
                'longitud': float(config_gps.longitud),
                'radio': config_gps.radio_permitido_metros
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'valido': False,
            'mensaje': 'Formato de datos inválido'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'valido': False,
            'mensaje': f'Error al validar GPS: {str(e)}'
        }, status=500)


@login_required
def capturar_gps_admin(request):
    """
    Página para que el admin capture y guarde la ubicación GPS de la oficina
    """
    # Verificar que sea admin o staff
    if not request.user.is_staff and not request.user.is_superuser:
        return redirect('dashboard')
    
    # Obtener configuración GPS actual
    config_gps = ConfiguracionGPS.obtener_configuracion_activa()
    
    context = {
        'config_gps': config_gps,
        'csrf_token': request.META.get('CSRF_COOKIE', '')
    }
    
    return render(request, 'asistencia/capturar_gps_admin.html', context)


@login_required
@require_POST
def guardar_gps_admin(request):
    """
    Endpoint para guardar las coordenadas GPS capturadas en el admin
    """
    import json
    
    # Verificar que sea admin o staff
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': 'Permiso denegado'
        }, status=403)
    
    try:
        datos = json.loads(request.body)
        latitud = datos.get('latitud')
        longitud = datos.get('longitud')
        nombre = datos.get('nombre', 'Oficina Principal')
        radio = datos.get('radio', 20)
        
        if not latitud or not longitud:
            return JsonResponse({
                'success': False,
                'message': 'Coordenadas GPS inválidas'
            }, status=400)
        
        # Desactivar configuraciones anteriores
        ConfiguracionGPS.objects.filter(activa=True).update(activa=False)
        
        # Crear o actualizar configuración
        config, created = ConfiguracionGPS.objects.get_or_create(
            nombre=nombre,
            defaults={
                'latitud': latitud,
                'longitud': longitud,
                'radio_permitido_metros': radio,
                'activa': True
            }
        )
        
        if not created:
            config.latitud = latitud
            config.longitud = longitud
            config.radio_permitido_metros = radio
            config.activa = True
            config.save()
        
        return JsonResponse({
            'success': True,
            'message': f'✅ Ubicación guardada exitosamente',
            'config': {
                'nombre': config.nombre,
                'latitud': float(config.latitud),
                'longitud': float(config.longitud),
                'radio': config.radio_permitido_metros
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Formato de datos inválido'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al guardar: {str(e)}'
        }, status=500)


@login_required
@require_POST
def marcar_evento(request, accion: str):
    usuario = request.user
    fecha = timezone.localdate()
    ahora = timezone.now()
    ip_empleado = obtener_ip_cliente(request)
    from django.conf import settings

    if not usuario.permite_remoto and not settings.DEBUG:
        ip_valida = IpOficinaAutorizada.objects.filter(ip_publica=ip_empleado, activa=True).exists()
        if not ip_valida:
            return JsonResponse(
                {
                    "status": "error",
                    "message": f"Acceso denegado. Tu IP actual ({ip_empleado}) no pertenece a la red autorizada.",
                },
                status=403,
            )

    registro, _ = RegistroAsistencia.objects.get_or_create(
        empleado=usuario,
        fecha=fecha,
        defaults={"ip_registro": ip_empleado},
    )

    latitud = request.POST.get("latitud")
    longitud = request.POST.get("longitud")

    if accion in ("entrada", "qr"):
        for validar in (
            validar_dia_sin_permiso_ni_feriado,
            validar_horario_para_marcacion,
            lambda u, f: validar_gps_para_marcacion(u, latitud, longitud),
        ):
            mensaje = validar(usuario, fecha)
            if mensaje:
                return JsonResponse({"status": "error", "message": mensaje}, status=400)

    if accion == "entrada":
        if registro.hora_entrada:
            return JsonResponse({"status": "error", "message": "Entrada ya registrada."}, status=400)
        registro.hora_entrada = ahora
        registro.ip_registro = ip_empleado
        
        if latitud and longitud:
            registro.latitud_entrada = latitud
            registro.longitud_entrada = longitud
            registro.precisión_entrada = request.POST.get("precisión", None)

        if usuario.horario and not usuario.horario.es_laborable(fecha):
            if obtener_recuperacion_pendiente(usuario):
                registro.estado = RegistroAsistencia.ESTADO_RECUPERACION
            else:
                registro.estado = RegistroAsistencia.ESTADO_FALTA
        elif usuario.horario and ahora.time() > usuario.horario.hora_entrada_con_tolerancia():
            registro.estado = RegistroAsistencia.ESTADO_TARDANZA
        else:
            registro.estado = RegistroAsistencia.ESTADO_A_TIEMPO
    elif accion == "inicio_almuerzo":
        if not registro.hora_entrada:
            return JsonResponse({"status": "error", "message": "Registra entrada primero."}, status=400)
        if registro.inicio_almuerzo:
            return JsonResponse({"status": "error", "message": "Almuerzo ya iniciado."}, status=400)
        registro.inicio_almuerzo = ahora
    elif accion == "fin_almuerzo":
        if not registro.inicio_almuerzo:
            return JsonResponse({"status": "error", "message": "Inicia almuerzo primero."}, status=400)
        if registro.fin_almuerzo:
            return JsonResponse({"status": "error", "message": "Almuerzo ya finalizado."}, status=400)
        registro.fin_almuerzo = ahora
    elif accion == "salida":
        if not registro.hora_entrada:
            return JsonResponse({"status": "error", "message": "Registra entrada primero."}, status=400)
        if registro.hora_salida:
            return JsonResponse({"status": "error", "message": "Salida ya registrada."}, status=400)
        
        actividad = request.POST.get("actividad", "").strip()
        if not actividad:
            return JsonResponse({"status": "error", "message": "Debes ingresar tu resumen de actividades del día para poder marcar la salida."}, status=400)

        registro.hora_salida = ahora
        
        if latitud and longitud:
            registro.latitud_salida = latitud
            registro.longitud_salida = longitud
            registro.precisión_salida = request.POST.get('precisión', None)
        
        registro.actividad_diaria = actividad
        registro.horas_netas_trabajadas = calcular_horas_netas(
            registro.hora_entrada,
            registro.hora_salida,
            registro.inicio_almuerzo,
            registro.fin_almuerzo,
        )

        if registro.estado == RegistroAsistencia.ESTADO_RECUPERACION:
            recuperacion = obtener_recuperacion_pendiente(usuario)
            if recuperacion:
                recuperacion.fecha_recuperacion = fecha
                recuperacion.horas_recuperadas = registro.horas_netas_trabajadas or timedelta(0)
                recuperacion.estado = RecuperacionDia.ESTADO_RECUPERADO
                recuperacion.save(update_fields=["fecha_recuperacion", "horas_recuperadas", "estado"])
    elif accion == "qr":
        # Validar código QR del usuario
        codigo_escaneado = request.POST.get("codigo", "").strip().upper()
        if codigo_escaneado != usuario.codigo_qr:
            return JsonResponse({"status": "error", "message": "Código QR inválido. El código no coincide con tu usuario."}, status=400)
            
        if registro.hora_entrada:
            return JsonResponse({"status": "error", "message": "Ya has marcado tu entrada hoy."}, status=400)
            
        registro.hora_entrada = ahora
        registro.ip_registro = ip_empleado
        if latitud and longitud:
            registro.latitud_entrada = latitud
            registro.longitud_entrada = longitud
            registro.precisión_entrada = request.POST.get("precisión", None)
        if usuario.horario and ahora.time() > usuario.horario.hora_entrada_con_tolerancia():
            registro.estado = RegistroAsistencia.ESTADO_TARDANZA
        else:
            registro.estado = RegistroAsistencia.ESTADO_A_TIEMPO
    else:
        return JsonResponse({"status": "error", "message": "Accion no valida."}, status=400)

    registro.save()
    return JsonResponse({"status": "ok", "message": f"Acción '{accion}' registrada exitosamente", "accion": accion})


@login_required
@require_POST
def escanear_qr_empleado(request):
    """
    Endpoint para RRHH: escanear código QR de empleado y validar/registrar asistencia.
    """
    # Verificar que el usuario sea RRHH o Admin
    if request.user.rol not in [CustomUser.ROL_RRHH, CustomUser.ROL_ADMIN]:
        return JsonResponse({"status": "error", "message": "No tienes permiso para realizar esta acción"}, status=403)
    
    try:
        datos = json.loads(request.body)
        codigo_qr = datos.get("codigo_qr", "").strip().upper()
        accion = datos.get("accion", "entrada")  # entrada, salida
        
        if not codigo_qr:
            return JsonResponse({"status": "error", "message": "Código QR requerido"}, status=400)
        
        # Buscar empleado por código QR
        empleado = CustomUser.objects.filter(codigo_qr=codigo_qr).first()
        if not empleado:
            return JsonResponse({"status": "error", "message": "Empleado no encontrado. Código QR inválido."}, status=404)
        
        # Crear o obtener registro de hoy
        fecha = timezone.localdate()
        ahora = timezone.now()
        registro, creado = RegistroAsistencia.objects.get_or_create(
            empleado=empleado,
            fecha=fecha,
            defaults={"ip_registro": obtener_ip_cliente(request)}
        )
        
        if accion == "entrada":
            if registro.hora_entrada:
                return JsonResponse(
                    {
                        "status": "warning",
                        "message": f"{empleado.get_full_name()} ya marcó entrada a las {registro.hora_entrada.strftime('%H:%M')}"
                    },
                    status=200
                )
            
            registro.hora_entrada = ahora
            registro.ip_registro = obtener_ip_cliente(request)
            
            # Determinar estado de entrada
            if empleado.horario:
                if not empleado.horario.es_laborable(fecha):
                    if obtener_recuperacion_pendiente(empleado):
                        registro.estado = RegistroAsistencia.ESTADO_RECUPERACION
                    else:
                        registro.estado = RegistroAsistencia.ESTADO_FALTA
                elif ahora.time() > empleado.horario.hora_entrada_con_tolerancia():
                    registro.estado = RegistroAsistencia.ESTADO_TARDANZA
                else:
                    registro.estado = RegistroAsistencia.ESTADO_A_TIEMPO
            else:
                registro.estado = RegistroAsistencia.ESTADO_A_TIEMPO
            
            registro.save()
            
            return JsonResponse(
                {
                    "status": "ok",
                    "message": f"✓ {empleado.get_full_name()} - Entrada marcada ({registro.get_estado_display()})",
                    "empleado": {
                        "nombre": empleado.get_full_name(),
                        "dni": empleado.dni,
                        "area": empleado.area.nombre if empleado.area else "Sin área",
                        "estado": registro.get_estado_display(),
                        "hora_entrada": registro.hora_entrada.strftime("%H:%M") if registro.hora_entrada else None
                    }
                },
                status=200
            )
        
        elif accion == "salida":
            if not registro.hora_entrada:
                return JsonResponse(
                    {"status": "error", "message": f"{empleado.get_full_name()} no ha marcado entrada hoy"},
                    status=400
                )
            
            if registro.hora_salida:
                return JsonResponse(
                    {
                        "status": "warning",
                        "message": f"{empleado.get_full_name()} ya marcó salida a las {registro.hora_salida.strftime('%H:%M')}"
                    },
                    status=200
                )
            
            registro.hora_salida = ahora
            
            # Calcular horas netas
            if registro.inicio_almuerzo and registro.fin_almuerzo:
                registro.horas_netas_trabajadas = calcular_horas_netas(
                    registro.hora_entrada,
                    registro.hora_salida,
                    registro.inicio_almuerzo,
                    registro.fin_almuerzo,
                )
            else:
                registro.horas_netas_trabajadas = registro.hora_salida - registro.hora_entrada
            
            registro.save()
            
            horas_trabajadas = formatear_duracion(registro.horas_netas_trabajadas)
            return JsonResponse(
                {
                    "status": "ok",
                    "message": f"✓ {empleado.get_full_name()} - Salida marcada ({horas_trabajadas}h)",
                    "empleado": {
                        "nombre": empleado.get_full_name(),
                        "dni": empleado.dni,
                        "area": empleado.area.nombre if empleado.area else "Sin área",
                        "hora_entrada": registro.hora_entrada.strftime("%H:%M") if registro.hora_entrada else None,
                        "hora_salida": registro.hora_salida.strftime("%H:%M") if registro.hora_salida else None,
                        "horas_trabajadas": horas_trabajadas
                    }
                },
                status=200
            )
        
        else:
            return JsonResponse({"status": "error", "message": "Acción no válida"}, status=400)
    
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "message": "Formato JSON inválido"}, status=400)
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Error: {str(e)}"}, status=500)
class ActividadesEmpleadosView(LoginRequiredMixin, AdminOrRRHHRequiredMixin, TemplateView):
    template_name = "asistencia/actividades_empleados.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        fecha_filtro = self.request.GET.get('fecha')
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        empleado_filtro = self.request.GET.get('empleado')
        
        registros = RegistroAsistencia.objects.filter(actividad_diaria__isnull=False).exclude(actividad_diaria='').select_related('empleado')
        
        if fecha_filtro:
            registros = registros.filter(fecha=fecha_filtro)
        if fecha_inicio:
            registros = registros.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            registros = registros.filter(fecha__lte=fecha_fin)
        if empleado_filtro:
            registros = registros.filter(empleado__id=empleado_filtro)
        
        registros = registros.order_by('-fecha')
        
        total_actividades = registros.count()
        total_empleados_reportaron = registros.values('empleado').distinct().count()
        
        context.update({
            'registros': registros[:100],
            'total_actividades': total_actividades,
            'total_empleados_reportaron': total_empleados_reportaron,
            'empleados': CustomUser.objects.filter(rol=CustomUser.ROL_EMPLEADO).order_by('last_name', 'first_name'),
        })
        return context

class AreasView(LoginRequiredMixin, AdminOrRRHHRequiredMixin, CreateView):
    template_name = "asistencia/areas.html"
    form_class = AreaForm
    success_url = "/areas/"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["areas"] = Area.objects.all().order_by("nombre")
        return context


class HorariosView(LoginRequiredMixin, AdminOrRRHHRequiredMixin, CreateView):
    template_name = "asistencia/horarios.html"
    form_class = HorarioForm
    success_url = "/horarios/"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["horarios"] = Horario.objects.all().order_by("nombre")
        context["feriados"] = DiaFeriado.objects.all().order_by("fecha")
        context["feriado_form"] = DiaFeriadoForm()
        context["permiso_form"] = AusenciaProgramadaForm()
        context["ausencias"] = AusenciaProgramada.objects.select_related("empleado").order_by("-fecha_inicio")
        context["recuperaciones"] = RecuperacionDia.objects.select_related("empleado").filter(
            estado=RecuperacionDia.ESTADO_PENDIENTE
        ).order_by("fecha_falta")
        return context


@login_required
@require_POST
def crear_ausencia_programada(request):
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
        return HttpResponse("No autorizado", status=403)

    form = AusenciaProgramadaForm(request.POST)
    if form.is_valid():
        permiso = form.save(commit=False)
        permiso.creada_por = request.user
        permiso.save()
    return redirect("horarios")


@login_required
@require_POST
def procesar_ausencia(request, ausencia_id: int):
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
        return HttpResponse("No autorizado", status=403)

    ausencia = get_object_or_404(AusenciaProgramada, pk=ausencia_id)
    accion = request.POST.get("accion")

    if accion == "aprobar":
        ausencia.estado = AusenciaProgramada.ESTADO_APROBADA
    elif accion == "rechazar":
        ausencia.estado = AusenciaProgramada.ESTADO_RECHAZADA
    else:
        return redirect("horarios")

    ausencia.procesada_por = request.user
    ausencia.save(update_fields=["estado", "procesada_por"])

    if ausencia.estado == AusenciaProgramada.ESTADO_APROBADA:
        RegistroAsistencia.objects.filter(
            empleado=ausencia.empleado,
            fecha__range=(ausencia.fecha_inicio, ausencia.fecha_fin),
            estado=RegistroAsistencia.ESTADO_FALTA,
        ).update(estado=RegistroAsistencia.ESTADO_PERMISO)

    return redirect("horarios")


@login_required
@require_POST
def actualizar_empleado_api(request, empleado_id: int):
    if request.user.rol not in (CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH, CustomUser.ROL_SUPERVISOR) and not request.user.is_staff:
        return JsonResponse({
            'success': False,
            'message': 'Permiso denegado'
        }, status=403)
    
    try:
        empleado = get_object_or_404(CustomUser, pk=empleado_id, is_active=True)
        datos = json.loads(request.body)
        if not isinstance(datos, dict):
            return JsonResponse({
                'success': False,
                'message': 'Formato de datos inválido'
            }, status=400)
        updated_fields = []
        
        if 'first_name' in datos:
            empleado.first_name = str(datos.get('first_name', '')).strip()
            updated_fields.append('first_name')
        if 'last_name' in datos:
            empleado.last_name = str(datos.get('last_name', '')).strip()
            updated_fields.append('last_name')
        if 'username' in datos:
            username = str(datos.get('username', '')).strip()
            if not username:
                return JsonResponse({
                    'success': False,
                    'message': 'El usuario es obligatorio'
                }, status=400)
            empleado.username = username
            updated_fields.append('username')
        if 'dni' in datos:
            dni = str(datos.get('dni', '')).strip()
            if not dni:
                return JsonResponse({
                    'success': False,
                    'message': 'El DNI es obligatorio'
                }, status=400)
            empleado.dni = dni
            updated_fields.append('dni')
        if 'permite_remoto' in datos:
            empleado.permite_remoto = bool(datos.get('permite_remoto'))
            updated_fields.append('permite_remoto')

        # Actualizar email
        if 'email' in datos:
            email = str(datos.get('email', '')).strip()
            if not email:
                return JsonResponse({
                    'success': False,
                    'message': 'El correo es obligatorio'
                }, status=400)
            validate_email(email)
            empleado.email = email
            updated_fields.append('email')
        
        # Actualizar área
        if 'area' in datos:
            if datos['area']:
                area = get_object_or_404(Area, pk=datos['area'])
                empleado.area = area
            else:
                empleado.area = None
            updated_fields.append('area')

        if 'horario' in datos:
            if datos['horario']:
                horario = get_object_or_404(Horario, pk=datos['horario'])
                empleado.horario = horario
            else:
                empleado.horario = None
            updated_fields.append('horario')
        
        # Actualizar rol
        if 'rol' in datos:
            rol = str(datos.get('rol', '')).strip().lower()
            if rol not in {CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH, CustomUser.ROL_SUPERVISOR, CustomUser.ROL_EMPLEADO, CustomUser.ROL_PPHH}:
                return JsonResponse({
                    'success': False,
                    'message': 'Rol inválido'
                }, status=400)
            if request.user.rol == CustomUser.ROL_RRHH and rol == CustomUser.ROL_ADMIN:
                return JsonResponse({
                    'success': False,
                    'message': 'RRHH no puede asignar rol administrador'
                }, status=403)
            empleado.rol = rol
            updated_fields.append('rol')
        
        if not updated_fields:
            return JsonResponse({
                'success': False,
                'message': 'No se enviaron campos para actualizar'
            }, status=400)

        empleado.full_clean()
        empleado.save(update_fields=updated_fields)
        
        return JsonResponse({
            'success': True,
            'message': 'Empleado actualizado correctamente'
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'JSON inválido'
        }, status=400)
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': mensajes_validation_error(e)
        }, status=400)
    except IntegrityError:
        return JsonResponse({
            'success': False,
            'message': 'Usuario o DNI ya registrado por otra persona'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def exportar_reporte_excel(request):
    """
    Exporta el reporte de asistencias filtrado a un archivo Excel (.xlsx).
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Verificar que el usuario sea Admin, RRHH o Supervisor
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_SUPERVISOR, CustomUser.ROL_RRHH]:
        return HttpResponse("No autorizado", status=403)

    # Obtener filtros de la URL (GET)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    empleado_id = request.GET.get('empleado')
    area_id = request.GET.get('area')

    registros = RegistroAsistencia.objects.all().select_related('empleado', 'empleado__area')

    if fecha_inicio:
        registros = registros.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha__lte=fecha_fin)
    if empleado_id:
        registros = registros.filter(empleado_id=empleado_id)
    if area_id:
        registros = registros.filter(empleado__area_id=area_id)

    registros = registros.order_by('-fecha', '-hora_entrada')

    # Crear libro y hoja de cálculo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Asistencias"

    # Mostrar líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11)
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_even = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Título principal
    ws.merge_cells('A1:I1')
    ws['A1'] = "REPORTE DE CONTROL DE ASISTENCIA"
    ws['A1'].font = font_title
    ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtítulo (Filtros aplicados)
    ws.merge_cells('A2:I2')
    sub_text = "Filtros aplicados: "
    filters = []
    if fecha_inicio:
        filters.append(f"Desde: {fecha_inicio}")
    if fecha_fin:
        filters.append(f"Hasta: {fecha_fin}")
    if empleado_id:
        try:
            emp = CustomUser.objects.get(id=empleado_id)
            filters.append(f"Empleado: {emp.get_full_name() or emp.username}")
        except CustomUser.DoesNotExist:
            pass
    if area_id:
        try:
            ar = Area.objects.get(id=area_id)
            filters.append(f"Área: {ar.nombre}")
        except Area.DoesNotExist:
            pass
    sub_text += ", ".join(filters) if filters else "Ninguno"
    ws['A2'] = sub_text
    ws['A2'].font = Font(name="Calibri", size=10, italic=True, color="595959")
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # Fila vacía
    ws.row_dimensions[3].height = 10

    # Cabeceras
    headers = [
        "Empleado", 
        "DNI", 
        "Área / Departamento", 
        "Fecha", 
        "Hora Entrada", 
        "Hora Salida", 
        "Estado", 
        "Horas Trabajadas",
        "Actividad Diaria"
    ]
    
    header_row = 4
    ws.row_dimensions[header_row].height = 25
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    # Filas de datos
    current_row = 5
    for reg in registros:
        ws.row_dimensions[current_row].height = 20
        
        nombre = reg.empleado.get_full_name() or reg.empleado.username
        dni = reg.empleado.dni or "-"
        area = reg.empleado.area.nombre if reg.empleado.area else "Sin Área"
        fecha_str = reg.fecha.strftime('%Y-%m-%d') if reg.fecha else "-"
        entrada_str = timezone.localtime(reg.hora_entrada).strftime('%I:%M:%S %p') if reg.hora_entrada else "-"
        salida_str = timezone.localtime(reg.hora_salida).strftime('%I:%M:%S %p') if reg.hora_salida else "-"
        
        estado_map = {
            'a_tiempo': 'Puntual',
            'tardanza': 'Tardanza',
            'falta': 'Falta'
        }
        estado = estado_map.get(reg.estado, reg.estado or "-")
        
        horas = formatear_duracion(reg.horas_netas_trabajadas) if reg.horas_netas_trabajadas else "-"
        actividad = reg.actividad_diaria or "-"

        row_data = [nombre, dni, area, fecha_str, entrada_str, salida_str, estado, horas, actividad]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_thin
            
            # Alineación específica
            if col_idx in [2, 4, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Color de fondo alterno
            if current_row % 2 == 0:
                cell.fill = fill_even

        current_row += 1

    # Auto-ajuste de ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Preparar respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Reporte_Asistencias_{timezone.localdate().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@require_POST
def crear_feriado(request):
    """
    Registra un nuevo día feriado.
    """
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
        return HttpResponse("No autorizado", status=403)
        
    form = DiaFeriadoForm(request.POST)
    if form.is_valid():
        form.save()
    return redirect("horarios")


@login_required
@require_POST
def eliminar_feriado(request, feriado_id: int):
    """
    Elimina un día feriado existente.
    """
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
        return HttpResponse("No autorizado", status=403)
        
    feriado = get_object_or_404(DiaFeriado, id=feriado_id)
    feriado.delete()
    return redirect("horarios")


@login_required
@require_POST
def actualizar_horario_api(request, horario_id: int):
    """
    Endpoint de API para actualizar un horario.
    """
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
        return JsonResponse({'success': False, 'message': 'Permiso denegado'}, status=403)
        
    try:
        horario = get_object_or_404(Horario, pk=horario_id)
        datos = json.loads(request.body)
        if not isinstance(datos, dict):
            return JsonResponse({'success': False, 'message': 'Formato de datos inválido'}, status=400)
        
        if 'nombre' in datos:
            horario.nombre = datos['nombre']
        if 'hora_entrada' in datos:
            horario.hora_entrada = datos['hora_entrada']
        if 'hora_salida' in datos:
            horario.hora_salida = datos['hora_salida']
        if 'tolerancia_minutos' in datos:
            horario.tolerancia_minutos = int(datos['tolerancia_minutos'])
            
        def parse_bool(value):
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                return value.strip().lower() in {'1', 'true', 'yes', 'si', 'sí', 'on'}
            return bool(value)

        for dia in ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']:
            if dia in datos:
                setattr(horario, dia, parse_bool(datos[dia]))

        horario.full_clean()
        horario.save()
        return JsonResponse({'success': True, 'message': 'Horario actualizado correctamente'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'JSON inválido'}, status=400)
    except ValidationError as e:
        return JsonResponse({'success': False, 'message': mensajes_validation_error(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
@require_POST
def actualizar_area_api(request, area_id: int):
    """
    Endpoint de API para actualizar un área.
    """
    if request.user.rol not in [CustomUser.ROL_ADMIN, CustomUser.ROL_RRHH]:
        return JsonResponse({'success': False, 'message': 'Permiso denegado'}, status=403)
        
    try:
        area = get_object_or_404(Area, pk=area_id)
        datos = json.loads(request.body)
        if not isinstance(datos, dict):
            return JsonResponse({'success': False, 'message': 'Formato de datos inválido'}, status=400)
        
        if 'nombre' in datos:
            area.nombre = datos['nombre']
        if 'descripcion' in datos:
            area.descripcion = datos['descripcion']

        area.full_clean()
        area.save()
        return JsonResponse({'success': True, 'message': 'Área actualizada correctamente'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'JSON inválido'}, status=400)
    except ValidationError as e:
        return JsonResponse({'success': False, 'message': mensajes_validation_error(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)



